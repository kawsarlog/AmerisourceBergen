from selenium import webdriver
import time
import csv
import openpyxl
import traceback
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

username = 'your_username'
password = 'ypur_password'

input_workbook_name = r"SAVINGS TEMPLATE.xlsx"
input_sheet_name = r"Sheet1"

output_workbook_name = r"output_example.xlsx"
output_sheet_name = r"Sheet1"

search_url = 'https://abcorder.amerisourcebergen.com/abcb2bstorefront/abcb2b/en/'
base_url = 'https://abcorder.amerisourcebergen.com'

# Read Excel
def read_excel_sheet(workbook_name, sheet_name):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(workbook_name)
        
        # Get the specified sheet
        sheet = workbook[sheet_name]
        
        # Initialize a list to store the data
        data = []
        
        # Iterate through each row in the sheet and append the cell values to the data list
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        
        # Add a print statement to show the retrieved data
        print(f"Data retrieved from sheet '{sheet_name}' in workbook '{workbook_name}':")
        print(f"Total {len(data)-1} rows found")
        # Return the 2D list containing the data
        return data
    
    except FileNotFoundError:
        print(f"Error: File '{workbook_name}' not found.")
        return None
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in workbook '{workbook_name}'.")
        return None
    except Exception as e:
        print(f"Error: An unexpected error occurred: {e}")
        return None

# Write on Excel
def append_to_excel_sheet(workbook_name, sheet_name, headline_data, data):
    try:
        new_sheet = True
        
        # Check if the workbook exists. If not, create a new workbook.
        try:
            workbook = openpyxl.load_workbook(workbook_name)
            new_sheet = False
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        
        # Delete the default "Sheet" if it exists
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        
        # Get the specified sheet or create a new one if it doesn't exist.
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]
        
        # If Sheet created Newly
        if new_sheet:
            #print('new_sheet', new_sheet)
            sheet.append(headline_data)
        
        # Append data to the sheet.
        if isinstance(data[0], list):  # Check if data is a 2D list
            for row in data:
                sheet.append(row)
        else:  # If data is a simple list, append it as a single row
            sheet.append(data)
        
        # Save the changes to the workbook.
        workbook.save(workbook_name)
        
        #print(f"Data appended to sheet '{sheet_name}' in workbook '{workbook_name}'.")
        print("Data added successfully.")
    
    except Exception as e:
        print(f"Error: An unexpected error occurred while appending data: {e}")

def driver_define():
    print(f"Chrome Browser Openning")
    driver_path = ChromeDriverManager().install()
    options = Options()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    s = Service(driver_path)
    driver = webdriver.Chrome(service=s, options =options)
    driver.maximize_window()
    return driver

def website_login(driver, username, password):

    driver.get(base_url)

    print("Wait Until Login page appears Element") 
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, '[name="j_username"]')))
    time.sleep(1)

    print(f"Username Sending")
    sent_username = driver.find_element(By.CSS_SELECTOR, '[name="j_username"]')
    sent_username.send_keys(username)
    time.sleep(0.2)

    print(f"Sending Password")
    sent_password = driver.find_element(By.CSS_SELECTOR, '[name="j_password"]')
    sent_password.send_keys(password)
    time.sleep(0.2)

    print(f"Click on Login")
    login_submit = driver.find_element(By.CSS_SELECTOR, '[name="uidPasswordLogon"]')
    login_submit.click()
    time.sleep(0.2)

    print("Wait Until Login Success") 
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, '[href="/abcb2bstorefront/abcb2b/en/logout"]')))

    print("----------Login Sucess----------")

def get_data(driver, working_data):
    product_title = working_data[0] #first column
    NDC_code = working_data[1] #secound column

    print(f"Working at: {NDC_code}")

    try:
        # Search and extract Data
        driver.get(search_url)
        src_box = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, '[id="js-site-search-input"]')))
        time.sleep(0.2)
        src_box.clear()
        time.sleep(0.1)
        NDC_code = str(NDC_code)
        NDC_code_replace = NDC_code.replace('-', '')
        src_box.send_keys(NDC_code_replace)
        time.sleep(0.2)
        search_click = driver.find_element(By.CSS_SELECTOR, '[id="js-site-search-btn-click"]')
        search_click.click()
        print('Searching...')
        print('Wait until Product Appears')
    except:
        # Search and extract Data
        driver.get(search_url)
        src_box = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, '[id="js-site-search-input"]')))
        time.sleep(0.2)
        src_box.clear()
        time.sleep(0.1)
        NDC_code = str(NDC_code)
        NDC_code_replace = NDC_code.replace('-', '')
        src_box.send_keys(NDC_code_replace)
        time.sleep(0.2)
        search_click = driver.find_element(By.CSS_SELECTOR, '[id="js-site-search-btn-click"]')
        search_click.click()
        print('Searching...')
        print('Wait until Product Appears')


    page_result_data = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[contains(@class,"styles-product-tile__name_") or contains(text(), "Expanding your search to All ABC Products")]'))).text
    
    # If has result then3 sec sleep
    if 'search to All ABC Products' not in page_result_data:
        time.sleep(3)
    else:
        time.sleep(1)
        
    try:
        AWP = driver.find_element(By.XPATH, '//div[@class="undefined undefined"]/following-sibling::div').text
    except:
        AWP = ''
    print(f'AWP: {AWP}')

    try:
        Accq_cost = driver.find_element(By.CSS_SELECTOR, '[class^="styles-product-tile__acqCost"]').text
    except:
        Accq_cost = ''
    print(f'Accq_cost: {Accq_cost}')

    data_to_append = [product_title, NDC_code, AWP, Accq_cost]
    return data_to_append

try:
    #Retrive Excel data
    input_datas = read_excel_sheet(input_workbook_name, input_sheet_name)
    headline_data = input_datas[0]
    working_datas = input_datas[1:]
except:
    print(traceback.format_exc())
    input('There is a problem with Openning the input excel file')

try:
    driver = driver_define()
except:
    print(traceback.format_exc())
    input("Chrome Driver opening Problem or Network issue")

try:
    website_login(driver, username, password)
except:
    print(traceback.format_exc())
    input("Login Problem at The Website")

for working_data in working_datas:
    try:
        data_to_append = get_data(driver, working_data)
        append_to_excel_sheet(output_workbook_name, output_sheet_name, headline_data, data_to_append)
    except:
        print(traceback.format_exc())

print('Program Completed')

driver.quit()
